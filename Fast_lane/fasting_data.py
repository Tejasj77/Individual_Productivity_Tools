import datetime
import random
from openpyxl import load_workbook

wb = load_workbook(filename='fasting.xlsx')
ws = wb.worksheets[0]
tday_whole = datetime.datetime.now()
print(tday_whole.year)
first_day = datetime.date(tday_whole.year,tday_whole.month,1)
#ind = first_day.weekday() + 2
ind = 6
print(ind)
totdays=0

if (tday_whole.month == 1 or tday_whole.month == 3 or tday_whole.month == 5 or tday_whole.month == 7 or tday_whole.month == 8 or tday_whole.month == 10 or tday_whole.month == 12):
    totdays = 31
elif tday_whole.month == 2:
    if (tday_whole.year % 4 == 0 and tday_whole.year % 100 !=0):
        totdays = 29
else:
    totdays = 30


book = load_workbook(filename='fasting.xlsx')
ws = book.worksheets[0]

date1 = 1
delta = 1

def first_week_print(variable,ending,date):
    for x in range(0,ending):
        ws.cell(row=2, column=variable).value = date
        date += 1
        variable+=1
    return date

flag = 0
for x in range(3,7):
    for y in range(2,9):
        if date1 ==1 and ind ==8:
            #y_coordinate = ind
            ws.cell(row=2,column=ind).value = date1
            date1 +=1
            flag = 1
        elif date1 ==1 and ind<8:
            ws.cell(row=2,column=ind).value = date1
            date1 +=1
            end = 8 - ind
            delta = ind + 1
            date1 = first_week_print(delta,end,date1)
            flag = 1
        else:
            if flag == 1:
                ws.cell(row=x,column=y-1).value = date1
                date1 +=1
                flag = 0
            else:
                print(y,date1)
                ws.cell(row=x, column=y).value = date1
                date1 += 1

book.save('fast1.xlsx')

"""
n = (totdays - tday_whole.date)/7
fast_list = []
for i in range(0,n):
    fast_list.append(random.randint((tday_whole.date+1+7*(n)),tday_whole.date+7*(n+1))
"""