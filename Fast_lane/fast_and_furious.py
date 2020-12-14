import datetime
import random
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.colors import YELLOW

#book = load_workbook(filename='fasting.xlsx')
#ws = book.worksheets[0]
book = Workbook()
ws = book.active


my_bold = Font(bold=True,color="00FF0000")



#totdays=0

#if (tday_whole.month == 1 or tday_whole.month == 3 or tday_whole.month == 5 or tday_whole.month == 7 or tday_whole.month == 8 or tday_whole.month == 10 or tday_whole.month == 12):
#    totdays = 31
#elif tday_whole.month == 2:
#    if (tday_whole.year % 4 == 0 and tday_whole.year % 100 !=0):
#        totdays = 29
#else:
#    totdays = 30



class today:
    def __init__(self):
        self.tday_whole = datetime.datetime.now()
        self.aaj_date = int(self.tday_whole.strftime("%d"))

        self.first_day = datetime.date(self.tday_whole.year, self.tday_whole.month, 1)
        print(self.first_day)
        self.ind = self.first_day.weekday() + 2
        self.totdays=0

        if (self.tday_whole.month == 1 or self.tday_whole.month == 3 or self.tday_whole.month == 5 or self.tday_whole.month == 7 or self.tday_whole.month == 8 or self.tday_whole.month == 10 or self.tday_whole.month == 12):
            self.totdays = 31
        elif self.tday_whole.month == 2:
            if (self.tday_whole.year % 4 == 0 and self.tday_whole.year % 100 !=0):
                self.totdays = 29
            else:
               self.totdays=28
        else:
            self.totdays = 30

    #    def rem_week_print(self,date,index,ending):
#        for x in range(0,index):
#            ws.cell(row=2,column=index).value = date
#            date +=1
#            index +=1
#        return date
    def rem_week_print1(self,date,last):
        self.date = date
        self.last = last
        print(self.list1)
        for x in range(3,8):
            for y in range(2,9):
                if self.date <= 31:
                    if int(self.date) in self.list1:
                        print("We are also here")
                        ws.cell(row=x,column=y).value = self.date
                        ws.cell(row=x, column=y).fill = PatternFill(fgColor=YELLOW,bgColor="FFC7CE", fill_type="solid")
                        ws.cell(row=x,column=y).font = my_bold
                    else:
                        ws.cell(row=x, column=y).value = self.date
                    self.date +=1

    def print_first_week(self):
        self.date1 = 1
        while (self.ind <= 8):
            if self.date1 ==1 and self.ind == 8:
                ws.cell(row=2, column=8).value = self.date1
                self.date1 += 1
                self.ind +=1
            elif self.ind<8:
                print(self.list1)
                if self.date1 in self.list1:
                    print("We are here")
                    ws.cell(row=2,column=self.ind).value = self.date1
                    ws.cell(row=2,column=self.ind).fill = PatternFill(fgColor=YELLOW,bgColor="FFC7CE",fill_type="solid")
                    ws.cell(row=2, column=self.ind).font = my_bold
                else:
                    ws.cell(row=2, column=self.ind).value = self.date1
                self.date1 +=1
                self.ind +=1
        self.rem_week_print1(self.date1,self.totdays)

    def randomize(self):
        self.n = 0
        self.list1 = []
        self.n = (self.totdays-self.aaj_date)/7
        x,y = 0,0
        for i in range(0,int(self.n)):
            x = (self.aaj_date+(7*(i)))
            y = (self.aaj_date+(7*(i+1)))
            self.list1.append(random.randint(x+1,y))
        print(self.list1)


first_obj = today()
first_obj.randomize()
first_obj.print_first_week()

def get_total_days(tarikh):
    if(tarikh.month ==1 or tarikh.month == 3 or tarikh.month == 5 or tarikh.month ==7 or tarikh.month ==8 or tarikh.month ==10 or tarikh.month ==12):
        return 31
    elif tarikh.month ==2:
        if (tarikh.year %4 ==0 and tarikh.year %100 !=0):
            return 29
        else:
            return 28
    else:
        return 30
"""

def produce_current():
    tday  = datetime.datetime.now()
    aaj_date = tday.strftime("%d")
    current_totdays = get_total_days(aaj_date)

inp1 = str(input("Press 1 if you want current month's fasting prediction or Press 2 if you want the future fasting prediction"))
if inp1 == '1':
    produce_current()
else:
    produce_future()

"""

book.save('new.xlsx')